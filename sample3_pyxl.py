#-*- coding: utf8 -*-
# python2.7.12
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('sample3.xlsx')
#print wb2.get_sheet_names()
for sheet in wb:
    print sheet.title
    '''
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = sheet.title
    str = sheet.title
    wb2.save(str+'.xlsx')
    '''
    wb3 = Workbook()
    str = sheet.title
    ws3 = wb3.active
    ws3.title = sheet.title
    for r in sheet.rows:
        row = [i.value for i in r]
        ws3.append(row)
    wb3.save(str+'.xlsx')